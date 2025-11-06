"""
Export Service

Provides Excel and PDF export capabilities for inventory reports.

Features:
- Excel export with formatting
- PDF generation with styling
- Multiple report types
- Custom templates
- Batch export
"""

import logging
import io
from decimal import Decimal
from typing import List, Dict, Optional, Any
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("openpyxl not installed. Excel export will not be available.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("reportlab not installed. PDF export will not be available.")


logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Service for exporting inventory reports to Excel format.
    """

    @staticmethod
    def create_workbook():
        """Create a new Excel workbook"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        return openpyxl.Workbook()

    @staticmethod
    def style_header_row(worksheet, row_num: int = 1):
        """Apply styling to header row"""
        if not EXCEL_AVAILABLE:
            return

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def auto_size_columns(worksheet):
        """Auto-size all columns based on content"""
        if not EXCEL_AVAILABLE:
            return

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def export_aging_report(aging_analyses: List, company_name: str) -> bytes:
        """
        Export aging analysis to Excel.

        Args:
            aging_analyses: List of ProductAgingAnalysis
            company_name: Company name for header

        Returns:
            Excel file bytes
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aging Analysis"

        # Title
        ws['A1'] = f"{company_name} - Inventory Aging Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            'Product Code', 'Product Name', 'Category', 'Quantity',
            'Value', 'Avg Age (Days)', 'Oldest (Days)', 'Velocity',
            'Days Since Movement', 'Risk', 'Recommendation'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)

        ExcelExportService.style_header_row(ws, 4)

        # Data
        row_num = 5
        for analysis in aging_analyses:
            ws.cell(row=row_num, column=1, value=analysis.product_code)
            ws.cell(row=row_num, column=2, value=analysis.product_name)
            ws.cell(row=row_num, column=3, value=analysis.category)
            ws.cell(row=row_num, column=4, value=float(analysis.total_quantity))
            ws.cell(row=row_num, column=5, value=float(analysis.total_value))
            ws.cell(row=row_num, column=6, value=analysis.average_age_days)
            ws.cell(row=row_num, column=7, value=analysis.oldest_stock_days)
            ws.cell(row=row_num, column=8, value=analysis.movement_velocity)
            ws.cell(row=row_num, column=9, value=analysis.days_since_last_movement)
            ws.cell(row=row_num, column=10, value=analysis.obsolescence_risk)
            ws.cell(row=row_num, column=11, value=analysis.recommended_action)

            # Color code by risk
            risk_colors = {
                'CRITICAL': 'FF0000',
                'HIGH': 'FFA500',
                'MEDIUM': 'FFFF00',
                'LOW': '00FF00'
            }
            if analysis.obsolescence_risk in risk_colors:
                ws.cell(row=row_num, column=10).fill = PatternFill(
                    start_color=risk_colors[analysis.obsolescence_risk],
                    end_color=risk_colors[analysis.obsolescence_risk],
                    fill_type="solid"
                )

            row_num += 1

        ExcelExportService.auto_size_columns(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_abc_analysis(abc_results: List, company_name: str) -> bytes:
        """
        Export ABC analysis to Excel.

        Args:
            abc_results: List of ABCClassification
            company_name: Company name

        Returns:
            Excel file bytes
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ABC Analysis"

        # Title
        ws['A1'] = f"{company_name} - ABC Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            'Product Code', 'Product Name', 'Annual Value',
            '% of Total', 'Cumulative %', 'ABC Class', 'Recommendation'
        ]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=4, column=col_num, value=header)

        ExcelExportService.style_header_row(ws, 4)

        # Data
        row_num = 5
        for result in abc_results:
            ws.cell(row=row_num, column=1, value=result.product_code)
            ws.cell(row=row_num, column=2, value=result.product_name)
            ws.cell(row=row_num, column=3, value=float(result.annual_consumption_value))
            ws.cell(row=row_num, column=4, value=float(result.percentage_of_total))
            ws.cell(row=row_num, column=5, value=float(result.cumulative_percentage))
            ws.cell(row=row_num, column=6, value=result.abc_class)
            ws.cell(row=row_num, column=7, value=result.recommendation)

            # Color code by class
            class_colors = {
                'A': '00FF00',
                'B': 'FFFF00',
                'C': 'FFA500'
            }
            if result.abc_class in class_colors:
                ws.cell(row=row_num, column=6).fill = PatternFill(
                    start_color=class_colors[result.abc_class],
                    end_color=class_colors[result.abc_class],
                    fill_type="solid"
                )

            row_num += 1

        ExcelExportService.auto_size_columns(ws)

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "ABC Analysis Summary"
        ws_summary['A1'].font = Font(size=14, bold=True)

        a_count = sum(1 for r in abc_results if r.abc_class == 'A')
        b_count = sum(1 for r in abc_results if r.abc_class == 'B')
        c_count = sum(1 for r in abc_results if r.abc_class == 'C')
        total = len(abc_results)

        ws_summary['A3'] = 'Class'
        ws_summary['B3'] = 'Count'
        ws_summary['C3'] = 'Percentage'

        ws_summary['A4'] = 'A'
        ws_summary['B4'] = a_count
        ws_summary['C4'] = f"{a_count/total*100:.1f}%" if total > 0 else "0%"

        ws_summary['A5'] = 'B'
        ws_summary['B5'] = b_count
        ws_summary['C5'] = f"{b_count/total*100:.1f}%" if total > 0 else "0%"

        ws_summary['A6'] = 'C'
        ws_summary['B6'] = c_count
        ws_summary['C6'] = f"{c_count/total*100:.1f}%" if total > 0 else "0%"

        ExcelExportService.style_header_row(ws_summary, 3)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_valuation_report(valuation_data: Dict, company_name: str) -> bytes:
        """
        Export valuation report to Excel.

        Args:
            valuation_data: Valuation report data
            company_name: Company name

        Returns:
            Excel file bytes
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valuation Report"

        # Title
        ws['A1'] = f"{company_name} - Inventory Valuation Report"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = ['Product', 'Warehouse', 'Quantity', 'Unit Cost', 'Total Value', 'Method']

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=4, column=col_num, value=header)

        ExcelExportService.style_header_row(ws, 4)

        # Data
        row_num = 5
        items = valuation_data.get('items', [])
        for item in items:
            ws.cell(row=row_num, column=1, value=item.get('product_code'))
            ws.cell(row=row_num, column=2, value=item.get('warehouse_code'))
            ws.cell(row=row_num, column=3, value=item.get('quantity'))
            ws.cell(row=row_num, column=4, value=item.get('unit_cost'))
            ws.cell(row=row_num, column=5, value=item.get('total_value'))
            ws.cell(row=row_num, column=6, value=item.get('method'))
            row_num += 1

        # Summary
        if 'summary' in valuation_data:
            summary = valuation_data['summary']
            ws.cell(row=row_num + 1, column=4, value="Total Value:")
            ws.cell(row=row_num + 1, column=4).font = Font(bold=True)
            ws.cell(row=row_num + 1, column=5, value=summary.get('total_value', 0))
            ws.cell(row=row_num + 1, column=5).font = Font(bold=True)

        ExcelExportService.auto_size_columns(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class PDFExportService:
    """
    Service for exporting inventory reports to PDF format.
    """

    @staticmethod
    def export_aging_report_pdf(aging_analyses: List, company_name: str) -> bytes:
        """
        Export aging analysis to PDF.

        Args:
            aging_analyses: List of ProductAgingAnalysis
            company_name: Company name

        Returns:
            PDF file bytes
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER
        )

        # Title
        title = Paragraph(f"{company_name}<br/>Inventory Aging Analysis", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        subtitle = Paragraph(
            f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))

        # Table data
        data = [[
            'Product', 'Category', 'Qty', 'Value',
            'Avg Age', 'Velocity', 'Risk'
        ]]

        for analysis in aging_analyses[:50]:  # Limit to 50 for PDF
            data.append([
                analysis.product_code,
                analysis.category[:15],
                f"{float(analysis.total_quantity):.0f}",
                f"${float(analysis.total_value):,.2f}",
                f"{analysis.average_age_days}d",
                analysis.movement_velocity,
                analysis.obsolescence_risk
            ])

        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_abc_analysis_pdf(abc_results: List, company_name: str) -> bytes:
        """
        Export ABC analysis to PDF.

        Args:
            abc_results: List of ABCClassification
            company_name: Company name

        Returns:
            PDF file bytes
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER
        )

        # Title
        title = Paragraph(f"{company_name}<br/>ABC Analysis Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        subtitle = Paragraph(
            f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))

        # Table data
        data = [[
            'Product', 'Annual Value', '% of Total',
            'Cumulative %', 'Class'
        ]]

        for result in abc_results[:50]:
            data.append([
                result.product_code,
                f"${float(result.annual_consumption_value):,.2f}",
                f"{float(result.percentage_of_total):.1f}%",
                f"{float(result.cumulative_percentage):.1f}%",
                result.abc_class
            ])

        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


class ExportService:
    """
    Unified export service that routes to Excel or PDF exporters.
    """

    @staticmethod
    def export_report(
        report_type: str,
        data: Any,
        format: str,
        company_name: str
    ) -> bytes:
        """
        Export a report in the specified format.

        Args:
            report_type: Type of report (aging, abc, valuation, etc.)
            data: Report data
            format: Export format ('excel' or 'pdf')
            company_name: Company name for header

        Returns:
            File bytes

        Raises:
            ValueError: If format or report type is invalid
            ImportError: If required library is not installed
        """
        format = format.lower()

        if format == 'excel':
            if report_type == 'aging':
                return ExcelExportService.export_aging_report(data, company_name)
            elif report_type == 'abc':
                return ExcelExportService.export_abc_analysis(data, company_name)
            elif report_type == 'valuation':
                return ExcelExportService.export_valuation_report(data, company_name)
            else:
                raise ValueError(f"Unknown report type: {report_type}")

        elif format == 'pdf':
            if report_type == 'aging':
                return PDFExportService.export_aging_report_pdf(data, company_name)
            elif report_type == 'abc':
                return PDFExportService.export_abc_analysis_pdf(data, company_name)
            else:
                raise ValueError(f"PDF export not implemented for report type: {report_type}")

        else:
            raise ValueError(f"Unknown format: {format}. Use 'excel' or 'pdf'")

    @staticmethod
    def get_content_type(format: str) -> str:
        """Get HTTP content type for format"""
        format = format.lower()
        if format == 'excel':
            return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format == 'pdf':
            return 'application/pdf'
        else:
            return 'application/octet-stream'

    @staticmethod
    def get_file_extension(format: str) -> str:
        """Get file extension for format"""
        format = format.lower()
        if format == 'excel':
            return '.xlsx'
        elif format == 'pdf':
            return '.pdf'
        else:
            return '.bin'
