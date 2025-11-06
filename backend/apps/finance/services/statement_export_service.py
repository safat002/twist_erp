"""
Financial Statement Export Service.

Exports financial statements to various formats:
- Excel (XLSX)
- PDF (via ReportLab)
- CSV
"""
import io
from decimal import Decimal
from datetime import date
from typing import Dict, List, Optional, BinaryIO
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class StatementExportService:
    """Service for exporting financial statements to various formats."""

    def __init__(self, statement_data: Dict, statement_type: str):
        """
        Initialize the export service.

        Args:
            statement_data: Statement data dictionary from service
            statement_type: 'trial_balance', 'balance_sheet', 'income_statement'
        """
        self.data = statement_data
        self.statement_type = statement_type

    def export_to_excel(self) -> io.BytesIO:
        """
        Export statement to Excel format.

        Returns:
            BytesIO buffer containing Excel file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        if self.statement_type == 'trial_balance':
            self._export_trial_balance_to_excel(worksheet)
        elif self.statement_type == 'balance_sheet':
            self._export_balance_sheet_to_excel(worksheet)
        elif self.statement_type == 'income_statement':
            self._export_income_statement_to_excel(worksheet)

        workbook.save(output)
        output.seek(0)
        return output

    def export_to_csv(self) -> io.StringIO:
        """
        Export statement to CSV format.

        Returns:
            StringIO buffer containing CSV data
        """
        output = io.StringIO()
        writer = csv.writer(output)

        if self.statement_type == 'trial_balance':
            self._export_trial_balance_to_csv(writer)
        elif self.statement_type == 'balance_sheet':
            self._export_balance_sheet_to_csv(writer)
        elif self.statement_type == 'income_statement':
            self._export_income_statement_to_csv(writer)

        output.seek(0)
        return output

    def _export_trial_balance_to_excel(self, worksheet):
        """Export trial balance to Excel worksheet."""
        # Title
        worksheet['A1'] = self.data['company'].name
        worksheet['A2'] = 'Trial Balance'
        worksheet['A3'] = f"As of {self.data['as_of_date']}"
        worksheet['A4'] = f"Currency: {self.data['currency']}"

        # Headers
        headers = ['Code', 'Account Name', 'Debit', 'Credit', 'Balance']
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Data
        row = 7
        for account in self.data['accounts']:
            worksheet.cell(row=row, column=1).value = account['code']
            worksheet.cell(row=row, column=2).value = '  ' * account['level'] + account['name']
            worksheet.cell(row=row, column=3).value = float(account['debit'])
            worksheet.cell(row=row, column=4).value = float(account['credit'])
            worksheet.cell(row=row, column=5).value = float(account['balance'])

            # Format parent accounts as bold
            if account.get('is_parent'):
                for col in range(1, 6):
                    worksheet.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        # Totals
        total_row = row + 1
        worksheet.cell(row=total_row, column=2).value = 'TOTAL'
        worksheet.cell(row=total_row, column=3).value = float(self.data['total_debit'])
        worksheet.cell(row=total_row, column=4).value = float(self.data['total_credit'])

        for col in range(1, 6):
            cell = worksheet.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='double'))

        # Format numbers
        for row_num in range(7, total_row + 1):
            for col in [3, 4, 5]:
                worksheet.cell(row=row_num, column=col).number_format = '#,##0.00'

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15

    def _export_balance_sheet_to_excel(self, worksheet):
        """Export balance sheet to Excel worksheet."""
        worksheet['A1'] = self.data['company'].name
        worksheet['A2'] = 'Statement of Financial Position (Balance Sheet)'
        worksheet['A3'] = f"As of {self.data['as_of_date']}"
        worksheet['A4'] = f"Currency: {self.data['currency']}"

        row = 6

        # ASSETS
        worksheet.cell(row=row, column=1).value = 'ASSETS'
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Current Assets
        worksheet.cell(row=row, column=1).value = 'Current Assets'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['assets']['current']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Current Assets'
        worksheet.cell(row=row, column=2).value = float(self.data['assets']['total_current'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Non-Current Assets
        worksheet.cell(row=row, column=1).value = 'Non-Current Assets'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['assets']['non_current']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Non-Current Assets'
        worksheet.cell(row=row, column=2).value = float(self.data['assets']['total_non_current'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Total Assets
        worksheet.cell(row=row, column=1).value = 'TOTAL ASSETS'
        worksheet.cell(row=row, column=2).value = float(self.data['total_assets'])
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).border = Border(top=Side(style='double'), bottom=Side(style='double'))
        row += 3

        # LIABILITIES
        worksheet.cell(row=row, column=1).value = 'LIABILITIES AND EQUITY'
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Current Liabilities
        worksheet.cell(row=row, column=1).value = 'Current Liabilities'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['liabilities']['current']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Current Liabilities'
        worksheet.cell(row=row, column=2).value = float(self.data['liabilities']['total_current'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Non-Current Liabilities
        if self.data['liabilities']['non_current']:
            worksheet.cell(row=row, column=1).value = 'Non-Current Liabilities'
            worksheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            for item in self.data['liabilities']['non_current']:
                worksheet.cell(row=row, column=1).value = '  ' + item['name']
                worksheet.cell(row=row, column=2).value = float(item['amount'])
                row += 1

            worksheet.cell(row=row, column=1).value = 'Total Non-Current Liabilities'
            worksheet.cell(row=row, column=2).value = float(self.data['liabilities']['total_non_current'])
            worksheet.cell(row=row, column=1).font = Font(bold=True)
            row += 2

        # Total Liabilities
        worksheet.cell(row=row, column=1).value = 'Total Liabilities'
        worksheet.cell(row=row, column=2).value = float(self.data['liabilities']['total'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Equity
        worksheet.cell(row=row, column=1).value = 'Equity'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['equity']['items']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Equity'
        worksheet.cell(row=row, column=2).value = float(self.data['equity']['total'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Total Liabilities and Equity
        worksheet.cell(row=row, column=1).value = 'TOTAL LIABILITIES AND EQUITY'
        worksheet.cell(row=row, column=2).value = float(self.data['total_liabilities_and_equity'])
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).border = Border(top=Side(style='double'), bottom=Side(style='double'))

        # Format numbers
        for row_num in range(1, row + 1):
            cell = worksheet.cell(row=row_num, column=2)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 20

    def _export_income_statement_to_excel(self, worksheet):
        """Export income statement to Excel worksheet."""
        worksheet['A1'] = self.data['company'].name
        worksheet['A2'] = 'Income Statement (Statement of Comprehensive Income)'
        worksheet['A3'] = f"Period: {self.data['period']['start']} to {self.data['period']['end']}"
        worksheet['A4'] = f"Currency: {self.data['currency']}"

        row = 6

        # Revenue
        worksheet.cell(row=row, column=1).value = 'Revenue'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['revenue']['items']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Revenue'
        worksheet.cell(row=row, column=2).value = float(self.data['revenue']['total'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Cost of Sales
        worksheet.cell(row=row, column=1).value = 'Cost of Sales'
        worksheet.cell(row=row, column=2).value = float(self.data['cost_of_sales'])
        row += 2

        # Gross Profit
        worksheet.cell(row=row, column=1).value = 'Gross Profit'
        worksheet.cell(row=row, column=2).value = float(self.data['gross_profit'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        worksheet.cell(row=row, column=2).font = Font(bold=True)
        worksheet.cell(row=row, column=2).border = Border(top=Side(style='thin'))
        row += 2

        # Operating Expenses
        worksheet.cell(row=row, column=1).value = 'Operating Expenses'
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for item in self.data['operating_expenses']['items']:
            worksheet.cell(row=row, column=1).value = '  ' + item['name']
            worksheet.cell(row=row, column=2).value = float(item['amount'])
            row += 1

        worksheet.cell(row=row, column=1).value = 'Total Operating Expenses'
        worksheet.cell(row=row, column=2).value = float(self.data['operating_expenses']['total'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Operating Profit
        worksheet.cell(row=row, column=1).value = 'Operating Profit'
        worksheet.cell(row=row, column=2).value = float(self.data['operating_profit'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        worksheet.cell(row=row, column=2).border = Border(top=Side(style='thin'))
        row += 2

        # Finance Costs
        if self.data['finance_costs'] != 0:
            worksheet.cell(row=row, column=1).value = 'Finance Costs'
            worksheet.cell(row=row, column=2).value = float(self.data['finance_costs'])
            row += 2

        # Profit Before Tax
        worksheet.cell(row=row, column=1).value = 'Profit Before Tax'
        worksheet.cell(row=row, column=2).value = float(self.data['profit_before_tax'])
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        # Tax Expense
        worksheet.cell(row=row, column=1).value = 'Tax Expense'
        worksheet.cell(row=row, column=2).value = float(self.data['tax_expense'])
        row += 2

        # Net Profit
        worksheet.cell(row=row, column=1).value = 'NET PROFIT'
        worksheet.cell(row=row, column=2).value = float(self.data['net_profit'])
        worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).font = Font(bold=True, size=12)
        worksheet.cell(row=row, column=2).border = Border(top=Side(style='double'), bottom=Side(style='double'))

        # Format numbers
        for row_num in range(1, row + 1):
            cell = worksheet.cell(row=row_num, column=2)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 20

    def _export_trial_balance_to_csv(self, writer):
        """Export trial balance to CSV."""
        writer.writerow([self.data['company'].name])
        writer.writerow(['Trial Balance'])
        writer.writerow([f"As of {self.data['as_of_date']}"])
        writer.writerow([f"Currency: {self.data['currency']}"])
        writer.writerow([])

        writer.writerow(['Code', 'Account Name', 'Debit', 'Credit', 'Balance'])

        for account in self.data['accounts']:
            writer.writerow([
                account['code'],
                '  ' * account['level'] + account['name'],
                str(account['debit']),
                str(account['credit']),
                str(account['balance'])
            ])

        writer.writerow([])
        writer.writerow(['', 'TOTAL', str(self.data['total_debit']), str(self.data['total_credit']), ''])

    def _export_balance_sheet_to_csv(self, writer):
        """Export balance sheet to CSV."""
        writer.writerow([self.data['company'].name])
        writer.writerow(['Balance Sheet'])
        writer.writerow([f"As of {self.data['as_of_date']}"])
        writer.writerow([f"Currency: {self.data['currency']}"])
        writer.writerow([])

        writer.writerow(['ASSETS', 'Amount'])
        writer.writerow(['Current Assets', ''])

        for item in self.data['assets']['current']:
            writer.writerow([item['name'], str(item['amount'])])

        writer.writerow(['Total Current Assets', str(self.data['assets']['total_current'])])
        writer.writerow([])

        writer.writerow(['Non-Current Assets', ''])
        for item in self.data['assets']['non_current']:
            writer.writerow([item['name'], str(item['amount'])])

        writer.writerow(['Total Non-Current Assets', str(self.data['assets']['total_non_current'])])
        writer.writerow([])
        writer.writerow(['TOTAL ASSETS', str(self.data['total_assets'])])

    def _export_income_statement_to_csv(self, writer):
        """Export income statement to CSV."""
        writer.writerow([self.data['company'].name])
        writer.writerow(['Income Statement'])
        writer.writerow([f"Period: {self.data['period']['start']} to {self.data['period']['end']}"])
        writer.writerow([f"Currency: {self.data['currency']}"])
        writer.writerow([])

        writer.writerow(['Revenue', 'Amount'])
        for item in self.data['revenue']['items']:
            writer.writerow([item['name'], str(item['amount'])])

        writer.writerow(['Total Revenue', str(self.data['revenue']['total'])])
        writer.writerow([])
        writer.writerow(['Cost of Sales', str(self.data['cost_of_sales'])])
        writer.writerow(['Gross Profit', str(self.data['gross_profit'])])
