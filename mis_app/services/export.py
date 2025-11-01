"""
Django Export Service
Advanced export functionality for reports and dashboards

Supports multiple formats with customization options
"""

import io
import logging
import pandas as pd
from typing import Dict, List, Optional, Any, Union
from django.http import HttpResponse
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import json

from ..models import ExportHistory, User
from ..utils import log_user_action

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting data in various formats
    """
    
    def __init__(self):
        self.supported_formats = {
            'csv': self._export_csv,
            'excel': self._export_excel,
            'json': self._export_json,
            'pdf': self._export_pdf,
            'html': self._export_html
        }
        
        self.max_export_rows = 100000  # Configurable limit
    
    def export_data(self, data: pd.DataFrame, format_type: str, 
                   filename: str = None, options: Dict = None, 
                   user: User = None) -> Dict[str, Any]:
        """
        Export data in specified format
        
        Args:
            data: DataFrame to export
            format_type: Export format (csv, excel, json, pdf, html)
            filename: Optional filename
            options: Export options
            user: User performing export
            
        Returns:
            Dictionary with export result
        """
        try:
            if format_type not in self.supported_formats:
                raise ValueError(f"Unsupported format: {format_type}")
            
            if len(data) > self.max_export_rows:
                raise ValueError(f"Export too large. Maximum {self.max_export_rows} rows allowed.")
            
            options = options or {}
            
            # Generate filename if not provided
            if not filename:
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_{timestamp}.{format_type}"
            
            # Call appropriate export method
            export_method = self.supported_formats[format_type]
            result = export_method(data, filename, options)
            
            # Log export activity
            if user:
                self._log_export(user, format_type, filename, len(data), result.get('size', 0))
            
            return {
                'success': True,
                'filename': filename,
                'format': format_type,
                'row_count': len(data),
                'column_count': len(data.columns),
                **result
            }
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _export_csv(self, data: pd.DataFrame, filename: str, options: Dict) -> Dict[str, Any]:
        """Export to CSV format"""
        output = io.StringIO()
        
        # CSV options
        delimiter = options.get('delimiter', ',')
        include_index = options.get('include_index', False)
        encoding = options.get('encoding', 'utf-8')
        
        data.to_csv(
            output, 
            sep=delimiter, 
            index=include_index,
            encoding=encoding
        )
        
        content = output.getvalue()
        output.close()
        
        # Create HTTP response
        response = HttpResponse(content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': content,
            'size': len(content.encode('utf-8'))
        }
    
    def _export_excel(self, data: pd.DataFrame, filename: str, options: Dict) -> Dict[str, Any]:
        """Export to Excel format with styling"""
        output = io.BytesIO()
        
        # Excel options
        sheet_name = options.get('sheet_name', 'Data')
        include_index = options.get('include_index', False)
        add_styling = options.get('add_styling', True)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=include_index)
            
            if add_styling:
                self._apply_excel_styling(writer.book[sheet_name], data)
        
        content = output.getvalue()
        output.close()
        
        # Create HTTP response
        response = HttpResponse(
            content, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': content,
            'size': len(content)
        }
    
    def _apply_excel_styling(self, worksheet, data: pd.DataFrame):
        """Apply styling to Excel worksheet"""
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply alternating row colors
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for row_num in range(2, len(data) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, len(data.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill
    
    def _export_json(self, data: pd.DataFrame, filename: str, options: Dict) -> Dict[str, Any]:
        """Export to JSON format"""
        # JSON options
        orient = options.get('orient', 'records')  # records, index, values, etc.
        indent = options.get('indent', 2)
        
        # Convert DataFrame to JSON
        if orient == 'records':
            json_data = data.to_dict('records')
        elif orient == 'index':
            json_data = data.to_dict('index')
        elif orient == 'values':
            json_data = data.values.tolist()
        else:
            json_data = data.to_dict(orient)
        
        # Handle NaN values
        json_str = json.dumps(json_data, indent=indent, default=str)
        
        # Create HTTP response
        response = HttpResponse(json_str, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': json_str,
            'size': len(json_str.encode('utf-8'))
        }
    
    def _export_pdf(self, data: pd.DataFrame, filename: str, options: Dict) -> Dict[str, Any]:
        """Export to PDF format"""
        output = io.BytesIO()
        
        # PDF options
        title = options.get('title', 'Data Export')
        page_size = options.get('page_size', letter)
        font_size = options.get('font_size', 9)
        include_summary = options.get('include_summary', True)
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=page_size)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
        
        # Add summary if requested
        if include_summary:
            summary_text = f"""
            <b>Export Summary:</b><br/>
            Rows: {len(data)}<br/>
            Columns: {len(data.columns)}<br/>
            Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
            """
            elements.append(Paragraph(summary_text, styles['Normal']))
            elements.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = []
        
        # Add headers
        headers = list(data.columns)
        table_data.append(headers)
        
        # Add data rows (limit for PDF)
        max_rows = options.get('max_rows', 1000)
        for idx, row in data.head(max_rows).iterrows():
            table_data.append([str(val) for val in row.values])
        
        # Create table
        table = Table(table_data)
        
        # Apply table styling
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), font_size - 1),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        content = output.getvalue()
        output.close()
        
        # Create HTTP response
        response = HttpResponse(content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': content,
            'size': len(content)
        }
    
    def _export_html(self, data: pd.DataFrame, filename: str, options: Dict) -> Dict[str, Any]:
        """Export to HTML format"""
        # HTML options
        title = options.get('title', 'Data Export')
        include_styling = options.get('include_styling', True)
        table_id = options.get('table_id', 'export-table')
        
        # Generate HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            {'<style>' + self._get_html_styles() + '</style>' if include_styling else ''}
        </head>
        <body>
            <div class="container">
                <h1>{title}</h1>
                <div class="summary">
                    <p><strong>Rows:</strong> {len(data)} | <strong>Columns:</strong> {len(data.columns)}</p>
                    <p><strong>Generated:</strong> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                {data.to_html(table_id=table_id, classes='data-table', escape=False)}
            </div>
        </body>
        </html>
        """
        
        # Create HTTP response
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': html_content,
            'size': len(html_content.encode('utf-8'))
        }
    
    def _get_html_styles(self) -> str:
        """Get CSS styles for HTML export"""
        return """
        body {
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 20px;
        }
        .summary {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .data-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        .data-table th {
            background-color: #007bff;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }
        .data-table td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        .data-table tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        .data-table tr:hover {
            background-color: #e3f2fd;
        }
        """
    
    def export_dashboard(self, dashboard_config: Dict, user: User, 
                        format_type: str = 'pdf', options: Dict = None) -> Dict[str, Any]:
        """Export dashboard to specified format"""
        try:
            options = options or {}
            
            if format_type == 'pdf':
                return self._export_dashboard_pdf(dashboard_config, options)
            elif format_type == 'html':
                return self._export_dashboard_html(dashboard_config, options)
            else:
                raise ValueError(f"Dashboard export format {format_type} not supported")
                
        except Exception as e:
            logger.error(f"Dashboard export error: {e}")
            return {'success': False, 'error': str(e)}
    
    def _export_dashboard_pdf(self, dashboard_config: Dict, options: Dict) -> Dict[str, Any]:
        """Export dashboard as PDF"""
        # This would require integration with a chart rendering library
        # For now, return a placeholder
        output = io.BytesIO()
        
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title = dashboard_config.get('title', 'Dashboard Export')
        
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 20))
        
        # Add dashboard metadata
        metadata_text = f"""
        <b>Dashboard Information:</b><br/>
        Title: {title}<br/>
        Widgets: {len(dashboard_config.get('widgets', []))}<br/>
        Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        elements.append(Paragraph(metadata_text, styles['Normal']))
        
        doc.build(elements)
        
        content = output.getvalue()
        output.close()
        
        filename = f"dashboard_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = HttpResponse(content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': content,
            'size': len(content),
            'filename': filename
        }
    
    def _export_dashboard_html(self, dashboard_config: Dict, options: Dict) -> Dict[str, Any]:
        """Export dashboard as HTML"""
        title = dashboard_config.get('title', 'Dashboard Export')
        widgets = dashboard_config.get('widgets', [])
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            <style>{self._get_dashboard_html_styles()}</style>
        </head>
        <body>
            <div class="dashboard-container">
                <header>
                    <h1>{title}</h1>
                    <p class="meta">Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </header>
                <div class="widgets-grid">
                    {''.join([self._render_widget_html(widget) for widget in widgets])}
                </div>
            </div>
        </body>
        </html>
        """
        
        filename = f"dashboard_{timezone.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return {
            'response': response,
            'content': html_content,
            'size': len(html_content.encode('utf-8')),
            'filename': filename
        }
    
    def _get_dashboard_html_styles(self) -> str:
        """Get CSS styles for dashboard HTML export"""
        return """
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f0f0f0;
        }
        .dashboard-container {
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            border-radius: 8px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }
        header h1 {
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }
        .meta {
            margin: 10px 0 0 0;
            opacity: 0.9;
        }
        .widgets-grid {
            padding: 30px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 20px;
        }
        .widget {
            background-color: #fafafa;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            padding: 20px;
            min-height: 200px;
        }
        .widget-title {
            font-size: 1.2em;
            font-weight: bold;
            color: #333;
            margin-bottom: 15px;
            border-bottom: 2px solid #667eea;
            padding-bottom: 5px;
        }
        """
    
    def _render_widget_html(self, widget: Dict) -> str:
        """Render individual widget as HTML"""
        title = widget.get('title', 'Widget')
        widget_type = widget.get('type', 'unknown')
        
        return f"""
        <div class="widget">
            <div class="widget-title">{title}</div>
            <div class="widget-content">
                <p><strong>Type:</strong> {widget_type}</p>
                <p><em>Widget data would be rendered here in a full implementation.</em></p>
            </div>
        </div>
        """
    
    def _log_export(self, user: User, format_type: str, filename: str, 
                   row_count: int, file_size: int):
        """Log export activity"""
        try:
            ExportHistory.objects.create(
                user=user,
                format_type=format_type,
                filename=filename,
                row_count=row_count,
                file_size=file_size
            )
            
            log_user_action(
                user, 'export_data', 'export', filename,
                f'Exported {row_count} rows to {format_type}',
                {
                    'format': format_type,
                    'filename': filename,
                    'row_count': row_count,
                    'file_size': file_size
                }
            )
        except Exception as e:
            logger.warning(f"Could not log export activity: {e}")
    
    def get_export_history(self, user: User, limit: int = 50) -> List[Dict]:
        """Get export history for user"""
        try:
            exports = ExportHistory.objects.filter(user=user).order_by('-created_at')[:limit]
            
            return [{
                'id': str(export.id),
                'format_type': export.format_type,
                'filename': export.filename,
                'row_count': export.row_count,
                'file_size': export.file_size,
                'created_at': export.created_at.isoformat()
            } for export in exports]
            
        except Exception as e:
            logger.error(f"Error getting export history: {e}")
            return []
    
    def get_supported_formats(self) -> List[Dict]:
        """Get list of supported export formats"""
        return [
            {
                'format': 'csv',
                'name': 'CSV (Comma Separated Values)',
                'description': 'Simple text format compatible with Excel and most tools',
                'max_rows': self.max_export_rows
            },
            {
                'format': 'excel',
                'name': 'Excel Workbook',
                'description': 'Microsoft Excel format with styling and formatting',
                'max_rows': self.max_export_rows
            },
            {
                'format': 'json',
                'name': 'JSON',
                'description': 'JavaScript Object Notation for web applications',
                'max_rows': self.max_export_rows
            },
            {
                'format': 'pdf',
                'name': 'PDF Document',
                'description': 'Portable Document Format for professional reports',
                'max_rows': 1000  # Smaller limit for PDF
            },
            {
                'format': 'html',
                'name': 'HTML Page',
                'description': 'Web page format with interactive features',
                'max_rows': self.max_export_rows
            }
        ]


# Global export service instance
export_service = ExportService()